"""
Comprehensive Multi-Phase Reporting System for Lust Rentals

Phase 1: Enhanced Excel Reports with per-property detail, expense matrix, and LLC summary.
Phase 2: Web Dashboard with interactive drill-down, trends, and property comparisons.
"""

import os
import logging
from typing import Dict, List, Optional
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.data_processing.processor import FinancialDataProcessor
from src.utils.config import load_config
from src.categorization.category_utils import normalize_category, get_display_name

logger = logging.getLogger(__name__)


class ComprehensiveReportGenerator:
    """Generates Phase 1 & 2 comprehensive reports."""

    def __init__(self, data_dir: Path = None):
        """Initialize the report generator."""
        config = load_config()
        self.data_dir = data_dir or config.data_dir
        self.processor = FinancialDataProcessor(data_dir=self.data_dir)
        self.reports_dir = os.path.join(self.data_dir, "reports")
        os.makedirs(self.reports_dir, exist_ok=True)

    def generate_phase1_excel(self, year: int, output_file: str = None) -> str:
        """
        PHASE 1: Generate comprehensive Excel report with multiple sheets.
        
        Sheets:
        1. LLC Summary (consolidated view)
        2. Expense Matrix (all properties × all categories)
        3. Per-property detailed sheets (one per property)
        
        Args:
            year: Tax year
            output_file: Output file path (optional)
            
        Returns:
            Path to generated Excel file
        """
        logger.info(f"Generating Phase 1 Excel report for {year}")
        
        # Load data
        result = self.processor.load_processed_data(year)
        income_df = result["income"]
        expenses_df = result["expenses"]
        
        # Normalize categories for display
        if not expenses_df.empty and 'category' in expenses_df.columns:
            expenses_df['category_display'] = expenses_df['category'].apply(
                lambda x: get_display_name(normalize_category(x))
            )
        
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Get all unique properties
        all_properties = sorted(set(
            list(income_df['property_name'].unique() if not income_df.empty else []) +
            list(expenses_df['property_name'].unique() if not expenses_df.empty else [])
        ))
        
        # 1. Create LLC Summary sheet
        ws_llc = wb.create_sheet(title="LLC Summary", index=0)
        self._write_llc_summary(ws_llc, income_df, expenses_df, year)
        
        # 2. Create Expense Matrix sheet
        ws_matrix = wb.create_sheet(title="Expense Matrix", index=1)
        self._write_expense_matrix(ws_matrix, expenses_df, all_properties)
        
        # 3. Create per-property detail sheets
        for prop in all_properties:
            # Excel sheet names are limited to 31 chars
            sheet_name = prop[:31].replace(":", "").replace("/", "-")
            ws_prop = wb.create_sheet(title=sheet_name)
            self._write_property_detail(ws_prop, prop, income_df, expenses_df)
        
        # Save
        if output_file is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = os.path.join(self.reports_dir, f"Lust_Rentals_{year}_Comprehensive_Report_{timestamp}.xlsx")
        
        wb.save(output_file)
        logger.info(f"Phase 1 report saved to {output_file}")
        return output_file

    def _write_llc_summary(self, ws, income_df, expenses_df, year):
        """Write LLC-level consolidated summary."""
        # Title
        ws['A1'] = f"LUST RENTALS LLC - {year} CONSOLIDATED SUMMARY"
        ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells('A1:D1')
        ws.row_dimensions[1].height = 30
        
        # Calculations
        total_income = income_df['amount'].sum() if not income_df.empty else 0
        total_expenses = expenses_df['amount'].sum() if not expenses_df.empty else 0
        net_income = total_income - total_expenses
        ratio = (total_expenses / total_income) if total_income > 0 else 0
        
        # Summary Metrics Table
        row = 3
        self._write_metric_row(ws, row, "TOTAL INCOME", total_income, fmt='$#,##0.00')
        self._write_metric_row(ws, row+1, "TOTAL EXPENSES", total_expenses, fmt='$#,##0.00')
        self._write_metric_row(ws, row+2, "NET INCOME", net_income, fmt='$#,##0.00', bold_val=True)
        self._write_metric_row(ws, row+3, "EXPENSE RATIO", ratio, fmt='0.0%')
        
        # Expense Breakdown Header
        row = 8
        ws[f'A{row}'] = "EXPENSE BREAKDOWN BY CATEGORY"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].border = Border(bottom=Side(style='medium'))
        ws.merge_cells(f'A{row}:B{row}')
        
        # Expense Breakdown Data
        row += 1
        ws[f'A{row}'] = "Category"
        ws[f'B{row}'] = "Amount"
        for cell in [ws[f'A{row}'], ws[f'B{row}']]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            cell.border = Border(bottom=Side(style='thin'))
        
        if not expenses_df.empty:
            category_totals = expenses_df.groupby('category_display')['amount'].sum().sort_values(ascending=False)
            for cat, amt in category_totals.items():
                row += 1
                ws[f'A{row}'] = cat
                ws[f'B{row}'] = amt
                ws[f'B{row}'].number_format = '$#,##0.00'
        
        # Formatting
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 20

    def _write_expense_matrix(self, ws, expenses_df, all_properties):
        """Write 2D matrix: Properties (rows) x Categories (cols)."""
        ws['A1'] = "EXPENSE MATRIX (Property x Category)"
        ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        ws.merge_cells('A1:Z1') # Arbitrary width
        
        if expenses_df.empty:
            ws['A3'] = "No expense data available."
            return

        # Pivot data
        pivot = expenses_df.pivot_table(
            index='property_name', 
            columns='category_display', 
            values='amount', 
            aggfunc='sum', 
            fill_value=0
        )
        
        # Ensure all properties exist (even if 0 expenses)
        for prop in all_properties:
            if prop not in pivot.index:
                pivot.loc[prop] = 0
        
        pivot = pivot.sort_index()
        
        # Add Totals
        pivot['TOTAL'] = pivot.sum(axis=1)
        pivot.loc['TOTAL'] = pivot.sum(axis=0)

        # Write Headers
        row = 3
        headers = ['Property'] + list(pivot.columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            cell.border = Border(bottom=Side(style='thin'), top=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))
            cell.alignment = Alignment(horizontal='center')
            
            # Auto-width roughly
            ws.column_dimensions[get_column_letter(col_idx)].width = max(len(str(header)) + 2, 12)

        # Write Data
        for r_idx, (idx_name, series) in enumerate(pivot.iterrows(), 1):
            curr_row = row + r_idx
            
            # Property Name Label
            ws.cell(row=curr_row, column=1, value=idx_name).font = Font(bold=(idx_name=='TOTAL'))
            
            for c_idx, val in enumerate(series, 2):
                cell = ws.cell(row=curr_row, column=c_idx, value=val)
                cell.number_format = '$#,##0.00;-$#,##0.00;-'
                
                # Style Total Row/Col
                if idx_name == 'TOTAL' or headers[c_idx-1] == 'TOTAL':
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        ws.column_dimensions['A'].width = 30
        ws.freeze_panes = 'B4'  # Freeze header and property column

    def _write_property_detail(self, ws, prop_name, income_df, expenses_df):
        """Write detailed sheet for a single property."""
        # Header
        ws['A1'] = f"PROPERTY DETAIL: {prop_name}"
        ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws.merge_cells('A1:E1')
        
        # Filter Data
        p_income = income_df[income_df['property_name'] == prop_name] if not income_df.empty else pd.DataFrame()
        p_expenses = expenses_df[expenses_df['property_name'] == prop_name] if not expenses_df.empty else pd.DataFrame()
        
        # Summary Section
        inc_total = p_income['amount'].sum() if not p_income.empty else 0
        exp_total = p_expenses['amount'].sum() if not p_expenses.empty else 0
        net = inc_total - exp_total
        
        row = 3
        self._write_metric_row(ws, row, "Total Income", inc_total, fmt='$#,##0.00')
        self._write_metric_row(ws, row+1, "Total Expenses", exp_total, fmt='$#,##0.00')
        self._write_metric_row(ws, row+2, "Net Income", net, fmt='$#,##0.00', bold_val=True)
        
        # Transaction Lists
        row += 4
        
        # Income Section
        ws[f'A{row}'] = "INCOME TRANSACTIONS"
        ws[f'A{row}'].font = Font(bold=True, color="4472C4")
        row += 1
        headers = ["Date", "Description", "Category", "Amount"]
        for idx, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=idx, value=h)
            c.font = Font(bold=True)
            c.border = Border(bottom=Side(style='thin'))
        
        if not p_income.empty:
            for _, trans in p_income.iterrows():
                row += 1
                ws.cell(row=row, column=1, value=trans.get('date', ''))
                ws.cell(row=row, column=2, value=trans.get('description', ''))
                ws.cell(row=row, column=3, value=trans.get('category', ''))
                c = ws.cell(row=row, column=4, value=trans['amount'])
                c.number_format = '$#,##0.00'
        
        # Expenses Section
        row += 3
        ws[f'A{row}'] = "EXPENSE TRANSACTIONS"
        ws[f'A{row}'].font = Font(bold=True, color="C00000")
        row += 1
        for idx, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=idx, value=h)
            c.font = Font(bold=True)
            c.border = Border(bottom=Side(style='thin'))

        if not p_expenses.empty:
            for _, trans in p_expenses.iterrows():
                row += 1
                ws.cell(row=row, column=1, value=trans.get('date', ''))
                ws.cell(row=row, column=2, value=trans.get('description', ''))
                ws.cell(row=row, column=3, value=trans.get('category_display', ''))
                c = ws.cell(row=row, column=4, value=trans['amount'])
                c.number_format = '$#,##0.00'

        # Widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 15

    def _write_metric_row(self, ws, row, label, value, fmt, bold_val=False):
        """Helper to write a key-value pair."""
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        ws[f'B{row}'].number_format = fmt
        if bold_val:
            ws[f'B{row}'].font = Font(bold=True)
